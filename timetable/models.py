import datetime
import json
import random
from typing import Iterable
from django.db import models
from django.contrib.auth.models import User
from rest_framework import status
from openpyxl import Workbook, styles
import tempfile
from account.models import Account

from core.models import Course, Staff, Venue
from timetable.qeueing import Queue, each
from timetable.generator import Generator as TimetableGenerator


def default_excluded_week_days():
    return [0]


def now_date():
    return datetime.datetime.now()


# Create your models here.
class Timetable(models.Model):
    title = models.CharField(max_length=50, null=True, blank=True)
    duration = models.IntegerField(
        default=5,
    )
    slot_per_day = models.IntegerField(default=4)
    is_current = models.BooleanField(default=False)
    courses = models.ManyToManyField(Course, blank=True)
    staffs = models.ManyToManyField(Staff, blank=True)
    venues = models.ManyToManyField(Venue, blank=True)
    created_on = models.DateTimeField(auto_created=True)
    last_updated = models.DateTimeField(auto_now=True)
    semester = models.IntegerField(default=1, choices=((1, "First"), (2, "Second")))
    session = models.CharField(default="2023/2024", max_length=15)

    start_date = models.DateField(default=datetime.datetime.now)
    end_date = models.DateField(default=datetime.datetime.now)
    excluded_week_days = models.JSONField(
        default=default_excluded_week_days, blank=True
    )
    excluded_days = models.JSONField(default=list, blank=True)

    @property
    def slot_size(self):
        return self.duration * self.slot_per_day

    def __str__(self):
        return self.title

    def days_count(self):
        total_days = (self.end_date - self.start_date).days
        diff = total_days
        days = 0
        current_date = self.start_date

        print(current_date)
        while current_date <= self.end_date:
            # Check if the current day is Monday (weekday() returns 0 for Monday)

            # print(self.excluded_days)
            # print(f"{str(current_date)} {self.excluded_days}, {str(current_date) in self.excluded_days or current_date.weekday() in self.excluded_week_days}")

            if (
                str(current_date) in self.excluded_days
                or current_date.weekday() in self.excluded_week_days
            ):
                days += 1
                diff -= 1
            # Move to the next day
            current_date += datetime.timedelta(days=1)

        print({"total_days": total_days, "excludes": days, "days": diff})

        return diff

    def timetable_days(self):
        total_days = (self.end_date - self.start_date).days
        diff = total_days
        days = 0
        current_date = self.start_date
        days_in = []

        print(current_date)
        while current_date <= self.end_date:
            # Check if the current day is Monday (weekday() returns 0 for Monday)

            # print(self.excluded_days)
            # print(f"{str(current_date)} {self.excluded_days}, {str(current_date) in self.excluded_days or current_date.weekday() in self.excluded_week_days}")

            if (
                str(current_date) in self.excluded_days
                or current_date.weekday() in self.excluded_week_days
            ):
                days += 1
                diff -= 1
            else:
                days_in.append(str(current_date))
            # Move to the next day

            current_date += datetime.timedelta(days=1)

        return days_in

    def dates(self):
        current_date = self.start_date
        mondays_count = 0

        while current_date <= self.end_date:
            # Check if the current day is Monday (weekday() returns 0 for Monday)
            if current_date.weekday() == 0:
                mondays_count += 1

            # Move to the next day
            current_date += datetime.timedelta(days=1)

        return mondays_count

    def export_excel(self, account_type = "student"):
        book = Workbook()
        main_sheet = book.active
        
        supervisors_sheet = book.create_sheet("supervisors")
        invigilators_sheet = book.create_sheet("Invigilators")

        supervisors_sheet.freeze_panes = "A2"
        supervisors_sheet["A1"] = "Course"
        supervisors_sheet["B1"] = "Staff Name"
        supervisors_sheet["C1"] = "Department"
        supervisors_sheet["D1"] = "Date"
        supervisors_sheet["E1"] = "Slot"

        supervisors_sheet.column_dimensions[f"A"].width = 30
        supervisors_sheet.column_dimensions[f"B"].width = 30
        supervisors_sheet.column_dimensions[f"C"].width = 12
        supervisors_sheet.column_dimensions[f"D"].width = 16
        supervisors_sheet.column_dimensions[f"E"].width = 5

        invigilators_sheet.freeze_panes = "A2"
        invigilators_sheet["A1"] = "Course"
        invigilators_sheet["B1"] = "Staff Name"
        invigilators_sheet["C1"] = "Department"
        invigilators_sheet["D1"] = "Date"
        invigilators_sheet["E1"] = "Slot"

        invigilators_sheet.column_dimensions[f"A"].width = 30
        invigilators_sheet.column_dimensions[f"B"].width = 30
        invigilators_sheet.column_dimensions[f"C"].width = 12
        invigilators_sheet.column_dimensions[f"D"].width = 16
        invigilators_sheet.column_dimensions[f"E"].width = 5

        dates = self.timetable_days()
        slots = TimetableSlot.objects.filter(timetable=self.pk).order_by("index")
        slots_queue = Queue(list(slots))

        row_index = 2
        column_indexes = "B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z".split(",")

        invigilator_index = 2

        if main_sheet:
            main_sheet[f"A1"] = f"Date"  # type: ignore
            main_sheet.column_dimensions["A"].width = 16  # type: ignore
            # main_sheet[f"A1"].alignment = styles.Alignment(wrap_text=True, ) # type: ignore

            for slot_ in range(self.slot_per_day):
                main_sheet[f"{column_indexes[slot_]}1"] = f"Slot {slot_+1}"  # type: ignore
                main_sheet.column_dimensions[f"{column_indexes[slot_]}"].width = 30  # type: ignore

            main_sheet.freeze_panes = "A2"  # type: ignore

        course_index = 2

        for date in dates:
            # Loop through the dates of the timetable
            if slots_queue.count() > 0:
                main_sheet[f"A{row_index}"] = date  # type: ignore

            try:
                for slot_index in range(self.slot_per_day):
                    slot = slots_queue.pop()
                    courses = slot.courses
                    buffer = ""

                    for course in courses:
                        venues_buffer = list(
                            map(lambda venue: venue.code, course.venues.all())
                        )

                        # Add course supervisor to shett
                        if account_type == "admin":
                            supervisors_sheet[f"A{course_index}"] = course.course.code
                            if course.supervisor:
                                
                                
                                supervisors_sheet[
                                    f"B{course_index}"
                                ] = course.supervisor.name
                                supervisors_sheet[
                                    f"C{course_index}"
                                ] = course.supervisor.department.code

                            supervisors_sheet[f"D{course_index}"] = date
                            supervisors_sheet[f"E{course_index}"] = slot_index + 1
                            
                        if account_type == "admin":
                            for staff in course.invigilators.all():
                                invigilators_sheet[
                                    f"A{invigilator_index}"
                                ] = course.course.code
                                invigilators_sheet[f"B{invigilator_index}"] = staff.name
                                invigilators_sheet[
                                    f"C{invigilator_index}"
                                ] = staff.department.code
                                invigilators_sheet[f"D{invigilator_index}"] = date
                                invigilators_sheet[f"E{invigilator_index}"] = slot_index + 1
                                invigilator_index = invigilator_index + 1

                        course_index = course_index + 1
                        buffer = f"{buffer} > {course.course.code} ({','.join(venues_buffer)})\n"

                    if main_sheet != None:
                        main_sheet[f"{column_indexes[slot_index]}{row_index}"] = buffer  # type: ignore
                        main_sheet[f"{column_indexes[slot_index]}{row_index}"].alignment = styles.Alignment(wrap_text=True)  # type: ignore

                row_index = row_index + 1

            except IndexError:
                break
        temp = tempfile.NamedTemporaryFile(prefix=str(random.randrange(1000000, 100000000)), suffix=".xlsx", delete=False)
        book.save(temp.name)
        return temp

    def generate(self):
        staffs: Queue[Staff] = Queue(list(self.staffs.all()))

        courses = self.courses.all()
        venues = self.venues.all()
        generator = TimetableGenerator(
            list(courses), self.slot_per_day, self.days_count()
        )
        generator.start()
        generator.assign_venues(list(venues))

        TimetableSlot.objects.filter(timetable=self.pk).delete()

        for day in range(self.days_count()):
            current_day = day + 1
            slots = generator.get_day_slots(current_day)

            for slot in slots:
                timetable_slot = TimetableSlot()
                timetable_slot.timetable = self
                timetable_slot.index = slot.index
                timetable_slot.day = slot.slot_day(self.slot_per_day)
                timetable_slot.save()

                for course in slot.courses:
                    slot_course = SlotCourse()
                    slot_course.slot = timetable_slot
                    slot_course.course = course
                    supervisor = staffs.magic_refilling(
                        lambda staff: staff.department == course.department
                    )
                    slot_course.supervisor = supervisor

                    slot_course.save()
                    if course in generator.course_assigned_venues:
                        slot_course.venues.set(generator.course_assigned_venues[course])
                        slot_course.save()

        return {
            "details": "Timetable generated",
            "ignored": len(generator.ignored_courses),
            "ignored_courses": map(
                lambda course: f"{course.title} {course.code}",
                generator.ignored_courses,
            ),
        }

    def auto_assign_invigilators(self,):
        """Tries to assign invigilators to all the courses in the timetable"""
        
        slots = TimetableSlot.objects.filter(timetable=self.pk)
        unasigned_slot_courses: set[SlotCourse] = set()
        for slot in slots:
            all_slot_courses = slot_courses([slot])

            

            for slot_course in all_slot_courses:
                try:
                    assign_invigilator(slot_course)
                except SlotCourse.NoStaffAvailable:
                    unasigned_slot_courses.add(slot_course)
            
        return (slots, unasigned_slot_courses)


class TimetableSlot(models.Model):
    timetable = models.ForeignKey(
        Timetable, on_delete=models.CASCADE, related_name="slots"
    )
    index = models.IntegerField()
    day = models.IntegerField(
        default=1,
    )

    class Meta:
        verbose_name_plural = "Slots"
        verbose_name = "Slot"

    def __str__(self):
        return f"{self.timetable} ({self.index})"

    @property
    def courses(self):
        return SlotCourse.objects.filter(slot=self.pk)

    @property
    def course_count(self):
        return SlotCourse.objects.filter(slot=self.pk).count()


class SlotCourse(models.Model):
    slot = models.ForeignKey(
        TimetableSlot, on_delete=models.CASCADE, related_name="slot_courses"
    )
    course = models.ForeignKey(Course, on_delete=models.CASCADE)
    supervisor = models.ForeignKey(
        Staff, null=True, on_delete=models.SET_NULL, related_name="supervising"
    )
    invigilators = models.ManyToManyField(
        Staff, blank=True, related_name="invigilating"
    )
    venues = models.ManyToManyField(Venue, blank=True)

    def complain_count(self):
        return Complain.objects.filter(slot_course=self.pk).count()

    class Meta:
        verbose_name_plural = "Slot Courses"
        verbose_name = "Slot Course"

    def __str__(self):
        return f"{self.course} {self.slot}"

    def auto_assign_venues(self, venues: set[Venue] | None = None):
        """Tries to assign venue automatically to this slot course

        Args:
            venues (set[Venue]): Pool of available venues to use

        Returns:
            boolean: True if the operation was successful
        """
        avialable_venues: list[Venue] = list()
        if not venues:
            # get all the venues for the timetable
            # and remove venues that are used by other adjacent slot  slotCourses
            # and slot courses in thesame slot as this slot course

            adjacent_slot_courses = SlotCourse.objects.filter(
                slot__day=self.slot.day,
                slot__index__in=[
                    self.slot.index - 1,
                    self.slot.index,
                    self.slot.index + 1,
                ],
            ).exclude(pk=self.pk)

            used_venues: set[Venue] = set()
            each(
                lambda sCourse: used_venues.update(sCourse.venues.all()),
                adjacent_slot_courses,
            )

            avialable_venues: list[Venue] = list(
                set(self.slot.timetable.venues.all()) - used_venues
            )

        else:
            avialable_venues = list(venues)

        # Sort the available venues by capacity from high to low and
        # convert it to queue, loop through each venue and pick venues
        # the best fit to contain the slot course course student count

        random.shuffle(avialable_venues)

        assigned_venues: list[Venue] = []
        count = 0
        done = False
        course = self.course

        ignored_venues: list[Venue] = list()

        for venue in avialable_venues:
            if (
                venue.capacity > course.student_count
                and venue.capacity <= course.student_count * 1.3
            ):
                assigned_venues = [venue]
                done = True
                break
            elif (count + venue.capacity) < (course.student_count * 1.3):
                assigned_venues.append(venue)
                count = count + venue.capacity
            else:
                ignored_venues.append(venue)

        if not done:
            sorted(ignored_venues, key=lambda venue: venue.capacity)

            for venue in ignored_venues:
                if count + venue.capacity < course.student_count * 1.3:
                    assigned_venues.append(venue)
                elif count >= course.student_count:
                    break

        if len(assigned_venues) == 0:
            assigned_venues.append(avialable_venues.pop())
        self.venues.set(assigned_venues)
        self.save()

        return True

    def auto_assign_invigilator(self, staffs: set[Staff] | None = None):
        """Tries to assign invigilators automatically based on the number
        of venues assigned to the slot course

        Returns:
            int: Status code that represent the success of the assignmnent
            200 : Means assignment was successfully
            406 : NO venue was assigned to slot course
            412 : No available staff

        """
        venues = set(self.venues.all())

        if len(venues) < 1:
            raise SlotCourse.NoVenueAssigned("No venue(s) was assigned to slot course")

        # Get slotCourse in thesame slot with this slotcourse and
        # extract all the staffs that are assgined as supervisors and
        # invigilators
        adjacent_slot_courses = SlotCourse.objects.filter(slot=self.slot.pk).exclude(
            pk=self.pk
        )
        unavailable_staffs: set[Staff] = set()

        for adjacent_slot_course in adjacent_slot_courses:
            unavailable_staffs.update(adjacent_slot_course.invigilators.all())

            if adjacent_slot_course.supervisor:
                unavailable_staffs.add(adjacent_slot_course.supervisor)

        # Get all slotcourses that are in adjacent slot to this course_slot slot
        # and extract all the staffs that are assgined as supervisors and
        # invigilators
        adjacent_slots_slot_courses = SlotCourse.objects.filter(
            slot__day=self.slot.day
        ).exclude(slot=self.slot)
        for adjacent_slot_course in adjacent_slots_slot_courses:
            unavailable_staffs.update(adjacent_slot_course.invigilators.all())

            if adjacent_slot_course.supervisor:
                unavailable_staffs.add(adjacent_slot_course.supervisor)

        # Get all staffs from the slot timetable that can invigilate
        # and filter then to staffs that are from the course department

        if staffs:
            course_department_staffs = filter(
                lambda staff: staff.department == self.course.department, staffs
            )

            if self.course.department:
                course_department_staffs = filter(
                    lambda staff: staff.department == self.course.department,
                    course_department_staffs,
                )
        else:
            course_department_staffs = self.slot.timetable.staffs.filter(
                can_invigilate=True,
            )

            if self.course.department:
                course_department_staffs = course_department_staffs.filter(
                    department=self.course.department.pk
                )

        available_staffs: set[Staff] = (
            set(course_department_staffs) - unavailable_staffs
        )

        # No available staff to assign
        if len(available_staffs) < 1:
            raise SlotCourse.NoStaffAvailable(
                "No available staff to invigilate slot course"
            )

        # Not enough staffs to assigned to number of venue
        elif len(available_staffs) < len(venues):
            self.invigilators.set(available_staffs)

        # No issue
        else:
            self.invigilators.set(list(available_staffs)[0 : len(venues)])

        self.save()

        return True

    def make_complain(self, message: str):
        complain = Complain(slot_course=self, message=message)
        complain.save()

    class NoStaffAvailable(Exception):
        pass

    class NoVenueAssigned(Exception):
        pass


class Complain(models.Model):
    LEVEL_CONFLICT = "level-conflict"
    OTHER_CONFLICT = "other-conflict"
    user = models.ForeignKey(User, on_delete=models.CASCADE, null=True, blank=True)
    slot_course = models.ForeignKey(
        SlotCourse, on_delete=models.CASCADE, related_name="complains"
    )
    resolved = models.BooleanField(default=False)
    complain_type = models.CharField(
        default=OTHER_CONFLICT,
        max_length=30,
        choices=(
            (LEVEL_CONFLICT, "Level course conflict"),
            (OTHER_CONFLICT, "Other confilct"),
        ),
    )
    related_slot_course = models.ForeignKey(
        SlotCourse, on_delete=models.CASCADE, null=True, blank=True
    )

    related_course = models.ForeignKey(
        Course, on_delete=models.CASCADE, null=True, default=None, blank=True
    )

    time = models.DateTimeField(auto_created=True, auto_now_add=True)

    def all_messages(
        self,
    ):
        return ComplainMessage.objects.filter(complain=self.pk)


class ComplainMessage(models.Model):
    user = models.ForeignKey(User, on_delete=models.CASCADE, null=True, blank=True)
    complain = models.ForeignKey(
        Complain, on_delete=models.CASCADE, related_name="messages"
    )
    reply_to = models.ForeignKey(
        "timetable.ComplainMessage",
        on_delete=models.CASCADE,
        related_name="replies",
        null=True,
        blank=True,
    )
    message = models.TextField(blank=True)
    time = models.DateTimeField(auto_created=True, auto_now_add=True)


def adjacent_slots(slot: TimetableSlot, include_self=False):
    indexes = (
        [slot.index - 1, slot.index + 1]
        if not include_self
        else [slot.index - 1, slot.index, slot.index + 1]
    )
    return TimetableSlot.objects.filter(
        timetable=slot.timetable.pk,
        day=slot.day,
        index__in=indexes,
    )


def nearby_slot_courses(slot_course: SlotCourse):
    return SlotCourse.objects.filter(slot=slot_course.slot.pk).exclude(
        pk=slot_course.pk
    )


def slot_courses(slots: Iterable[TimetableSlot]):
    return SlotCourse.objects.filter(slot__in=map(lambda slot: slot.pk, slots))


def assign_invigilator(
    slot_course: SlotCourse, staffs: Iterable[Staff] | None = None
) -> Iterable[Staff]:
    """Tries to assign invigilators to all the venues used in the course slot"""
    venues = slot_course.venues
    if venues.count() < 1:
        return []

    # course has venues

    available_staffs: set[Staff] = set()

    if staffs and slot_course.course.department:
        available_staffs.update(
            filter(
                lambda staff: staff.department == slot_course.course.department, staffs
            )
        )
    elif staffs:
        available_staffs.update(staffs)
    elif staffs == None and slot_course.course.department:
        available_staffs.update(
            slot_course.slot.timetable.staffs.filter(
                department=slot_course.course.department.pk
            )
        )
    else:
        available_staffs.update(slot_course.slot.timetable.staffs.all())

    if not staffs:
        # NO staffs was provided to filter out staffs that are already in the previous slots

        busy_staffs: set[Staff] = set()

        # Loop over each adjacent slot

        in_range_slots = adjacent_slots(slot_course.slot, include_self=True)
        for adjacent_slot_course in slot_courses(in_range_slots).exclude(
            pk=slot_course.pk
        ):
            busy_staffs.update(adjacent_slot_course.invigilators.all())
            if adjacent_slot_course.supervisor:
                busy_staffs.add(adjacent_slot_course.supervisor)

        available_staffs = available_staffs - busy_staffs

    if len(available_staffs) == 0:
        raise SlotCourse.NoStaffAvailable("No enough staff to assign as invigilators")

    # Available staffs is insufficient to assign to venues
    if len(available_staffs) < venues.count():
        slot_course.invigilators.set(available_staffs)
        return available_staffs

    selected_Staffs = map(lambda index: available_staffs.pop(), range(venues.count()))
    slot_course.invigilators.set(selected_Staffs)

    return selected_Staffs
